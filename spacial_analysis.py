import openpyxl
import math

file = 'Wells.xlsx'

maxDistance = 0.0005 #degrees

wb = openpyxl.load_workbook(file)
sheet = wb.get_sheet_by_name('Sheet0')

#dump the excel sheet into a dictionary
wells = {}
for row in range(2, sheet.max_row + 1):
    api = sheet['N' + str(row)].value
    wellname = sheet['B' + str(row)].value
    wellnumber = sheet['C' + str(row)].value
    operator = sheet['M' + str(row)].value
    county = sheet['A' + str(row)].value
    zone = sheet['G' + str(row)].value
    firstprod = sheet['H' + str(row)].value
    lat = sheet['R' + str(row)].value
    long = sheet['S' + str(row)].value
    wells[api] = dict(wellname=wellname, wellnumber=wellnumber, operator=operator, county=county, zone=zone, firstprod=firstprod, lat=lat, long=long, neighbors={})

#remove wells with no lat longs
for api in list(wells):
    if type(wells[api]['lat']) is str:
        del wells[api]

output = openpyxl.Workbook()
sheet = output.get_sheet_by_name('Sheet')
sheet['A1'] = 'Well Name'
sheet['B1'] = 'Well Number'
sheet['C1'] = 'API'
sheet['D1'] = 'Operator'
sheet['E1'] = 'County'
sheet['F1'] = 'Zone'
sheet['G1'] = 'First Prod'
sheet['H1'] = 'Latitude'
sheet['I1'] = 'Longitude'
sheet['J1'] = 'Neighbor Count'
sheet['K1'] = 'Neighbor Well Name'
sheet['L1'] = 'Neighbor Well Number'
sheet['M1'] = 'Neighbor API'
sheet['N1'] = 'Neighbor Operator'
sheet['O1'] = 'Neighbor County'
sheet['P1'] = 'Neighbor Zone'
sheet['Q1'] = 'Neighbor First Prod'
sheet['R1'] = 'Neighbor Latitude'
sheet['S1'] = 'Neighbor Longitude'

distances = {}
#iterate over all values and check if distance between two points is within maxDistance
for api in list(wells):
    for otherapi in list(wells):
        if otherapi == api:
            continue
        distance = math.sqrt((wells[api]['lat'] - wells[otherapi]['lat'])**2 + (wells[api]['long'] - wells[otherapi]['long'])**2)
        if distance < maxDistance:
            neighborwellname = wells[otherapi]['wellname']
            neighborwellnumber = wells[otherapi]['wellnumber']
            neighboroperator = wells[otherapi]['operator']
            neighborcounty = wells[otherapi]['county']
            neighborzone = wells[otherapi]['zone']
            neighborfirstprod = wells[otherapi]['firstprod']
            neighborlat = wells[otherapi]['lat']
            neighborlong = wells[otherapi]['long']
            wells[api]['neighbors'][otherapi] = dict(wellname=neighborwellname, wellnumber=neighborwellnumber, operator=neighboroperator, county=neighborcounty, zone=neighborzone, firstprod=neighborfirstprod, lat=neighborlat, long=neighborlong, distance=distance)

wellsreviewed = set()
rowNum = 2
for api in wells:
    if api in wellsreviewed:
        continue
    neighborwells = wells[api]['neighbors']
    sheet.cell(row=rowNum, column=1).value = wells[api]['wellname']
    sheet.cell(row=rowNum, column=2).value = wells[api]['wellnumber']
    sheet.cell(row=rowNum, column=3).value = api
    sheet.cell(row=rowNum, column=4).value = wells[api]['operator']
    sheet.cell(row=rowNum, column=5).value = wells[api]['county']
    sheet.cell(row=rowNum, column=6).value = wells[api]['zone']
    sheet.cell(row=rowNum, column=7).value = wells[api]['firstprod']
    sheet.cell(row=rowNum, column=8).value = wells[api]['lat']
    sheet.cell(row=rowNum, column=9).value = wells[api]['long']
    sheet.cell(row=rowNum, column=10).value = len(neighborwells)
    if len(neighborwells) == 0:
        rowNum = rowNum + 1
        continue
    for neighborapi in neighborwells:
        sheet.cell(row=rowNum, column=11).value = neighborwells[neighborapi]['wellname']
        sheet.cell(row=rowNum, column=12).value = neighborwells[neighborapi]['wellnumber']
        sheet.cell(row=rowNum, column=13).value = neighborapi
        sheet.cell(row=rowNum, column=14).value = wells[neighborapi]['operator']
        sheet.cell(row=rowNum, column=15).value = wells[neighborapi]['county']
        sheet.cell(row=rowNum, column=16).value = wells[neighborapi]['zone']
        sheet.cell(row=rowNum, column=17).value = wells[neighborapi]['firstprod']
        sheet.cell(row=rowNum, column=18).value = neighborwells[neighborapi]['lat']
        sheet.cell(row=rowNum, column=19).value = neighborwells[neighborapi]['long']
        rowNum = rowNum + 1
        wellsreviewed.add(neighborapi)
    wellsreviewed.add(api)

output.save('neighbors.xlsx')